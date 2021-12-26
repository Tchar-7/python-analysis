import tkinter as tk
from tkinter import StringVar, ttk, messagebox
import openpyxl
import joblib
import pandas as pd
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis as LDA
from sklearn.decomposition import PCA
import numpy as np
from sklearn.preprocessing import StandardScaler


class stone:
    def __init__(self, r, name1, lv1, name2, lv2, hole):
        self.rarity = r
        self.name1 = name1
        if self.name1 == "无":
            self.lv1 = '0'
        else:
            self.lv1 = lv1
        self.name2 = name2
        if self.name2 == "无":
            self.lv2 = '0'
        else:
            self.lv2 = lv2
        self.hole = hole

    def judge(self):
        estimator = joblib.load('cls.model')
        skill = pd.read_excel("skill.xlsx")
        stone_old= pd.read_excel("半监督.xlsx")
        stone_old=stone_old.iloc[:,:6]
        # print(stone_old.head())

        stone = pd.DataFrame([[self.rarity, self.name1, self.lv1, self.name2, self.lv2, self.hole]], columns=['等级', '一技能', 'Lv', '二技能', 'Lv1', '孔'])
        skillArray = skill["LookUpName"].values
        
        stone = pd.concat((stone_old, stone), axis=0)
        
        # 处理缺失值
        stone.fillna({"二技能":"无","Lv1":0},inplace=True)

        #处理技能,将技能名称用数字替换
        skillArray = skill["LookUpName"].values
        stone['一技能'].replace(skillArray,list(range(1,113)),inplace=True)
        stone['二技能'].replace(skillArray,list(range(1,113)),inplace=True)

        lda = LDA(n_components=1)
        pca = PCA(n_components=1)
        skill1=stone[['一技能','Lv']]
        re_1=lda.fit_transform(skill1,stone['一技能'])
        skill2=stone[['二技能','Lv1']]
        re_2=lda.fit_transform(skill2,stone['二技能'])
        skill = np.concatenate((re_1, re_2), axis=1)
        re_skill = pca.fit_transform(skill)
        stone=stone.reset_index()
        temp= pd.DataFrame(re_skill, columns=['temp'])
        temp=temp.reset_index()
        df = pd.concat((stone,temp), axis=1)
        print(df)
        df.drop(['一技能','Lv','二技能','Lv1'],axis = 1,inplace=True)
        stone = df
        #数据归一化
        kong=stone['孔'].values.reshape(-1,1)
        temp=stone['temp'].values.reshape(-1,1)
        rank=stone['等级'].values.reshape(-1,1)

        scaler = StandardScaler(copy=False)
        stone['孔']=scaler.fit_transform(kong)
        stone['temp']=scaler.fit_transform(temp)
        stone['等级']=scaler.fit_transform(rank)

        data=stone.iloc[[-1]]
        del data['index']
        print(data)
        tk.messagebox.showinfo('结果', estimator.predict(data))


class MH_GUI:
    def __init__(self):
        self.top = tk.Tk()
        self.top.geometry("500x300")
        self.top.resizable(0, 0)
        self.top.title("Monster Hunter Stone")
        self.name_list = []
        self.hole_list = []
        self.rarit_list = []
        self.read_name_list('./skill.xlsx', './stone.xlsx')
        self.r_list = ['0', '1', '2', '3']
        self.init_combobox()
        self.init_lables()
        self.init_button()
        self.top.mainloop()

    def init_button(self):
        self.button = tk.Button(self.top, text="评判", command=lambda: self.action()).place(x=240, y=250)

    def action(self):
        TheStone = stone(self.value6.get(), self.value1.get(), self.value2.get(), self.value3.get(), self.value4.get(), self.value5.get())
        TheStone.judge()

    def init_lables(self):
        self.label1 = tk.Label(self.top, text="技能1:").place(x=150, y=10)
        self.label2 = tk.Label(self.top, text="Lv:").place(x=150, y=50)
        self.label3 = tk.Label(self.top, text="技能2:").place(x=150, y=90)
        self.label4 = tk.Label(self.top, text="Lv:").place(x=150, y=130)
        self.label5 = tk.Label(self.top, text="孔位").place(x=150, y=170)
        self.label5 = tk.Label(self.top, text="稀有度").place(x=150, y=210)

    def init_combobox(self):
        self.value1 = StringVar()
        self.value1.set("请选择技能1")
        self.value2 = StringVar()
        self.value2.set("选择其等级")
        self.value3 = StringVar()
        self.value3.set("请选择技能2")
        self.value4 = StringVar()
        self.value4.set("选择其等级")
        self.value5 = StringVar()
        self.value5.set("请选择其孔位")
        self.value6 = StringVar()
        self.value6.set("请选择其稀有度")
        self.combobox1 = ttk.Combobox(self.top, textvariable=self.value1, height=10, width=20, values=self.name_list).place(x=200, y=10)
        self.combobox2 = ttk.Combobox(self.top, textvariable=self.value2, height=len(self.r_list), width=20, values=self.r_list).place(x=200, y=50)
        self.combobox3 = ttk.Combobox(self.top, textvariable=self.value3, height=10, width=20, values=self.name_list).place(x=200, y=90)
        self.combobox4 = ttk.Combobox(self.top, textvariable=self.value4, height=len(self.r_list), width=20, values=self.r_list).place(x=200, y=130)
        self.combobox5 = ttk.Combobox(self.top, textvariable=self.value5, height=10, width=20, values=self.hole_list).place(x=200, y=170)
        self.combobox6 = ttk.Combobox(self.top, textvariable=self.value6, height=10, width=20, values=self.rarit_list).place(x=200, y=210)

    def read_name_list(self, data_res, data_res1):
        readbook = openpyxl.load_workbook(data_res)
        sheet = readbook['skill']
        rows = sheet.max_row
        for i in range(2, rows+1):
            data = sheet.cell(i, 1).value
            self.name_list.append(data)
        readbook = openpyxl.load_workbook(data_res1)
        sheet = readbook['マカ錬金 幽玄 -ZH']
        rows = sheet.max_row
        for i in range(2, rows+1):
            data = sheet.cell(i, 6).value   # 第六列为孔位
            self.hole_list.append(data)
        self.hole_list = list(set(self.hole_list))
        for i in range(2, rows+1):
            data = sheet.cell(i, 1).value   # 第一列为稀有度
            self.rarit_list.append(data)
        self.rarit_list = list(set(self.rarit_list))


mh = MH_GUI()